"""Exportación profesional a un libro Excel de seis hojas."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .reconciliacion import ResultadoConciliacion


NOMBRES_HOJAS = (
    "Resumen",
    "Coincidencias",
    "Solo_en_A",
    "Solo_en_B",
    "Diferencias",
    "Duplicados",
)

COLOR_FONDO = "12202D"
COLOR_ENCABEZADO = "1D3448"
COLOR_ACENTO = "3C8DBC"
COLOR_TEXTO = "F2F6F8"
COLOR_BORDE = "496274"
COLOR_ALERTA = "6B3940"
COLOR_FILA_ALTERNA = "EAF1F5"
COLOR_BORDE_DATOS = "D6E0E7"


def _ancho_columna(hoja, indice: int) -> float:
    valores = [
        "" if celda.value is None else str(celda.value)
        for celda in list(hoja.columns)[indice - 1][:200]
    ]
    longitud = max((len(valor) for valor in valores), default=10)
    return min(max(longitud + 2, 12), 42)


def _proteger_texto_formula(hoja) -> None:
    """Conserva como texto los valores originales que empiezan con '='."""

    for fila in hoja.iter_rows():
        for celda in fila:
            if celda.data_type == "f":
                celda.data_type = "s"


def _formatear_hoja(hoja) -> None:
    hoja.freeze_panes = "A2"
    hoja.sheet_view.showGridLines = False
    hoja.row_dimensions[1].height = 28
    hoja.auto_filter.ref = hoja.dimensions if hoja.max_row >= 2 else None

    relleno = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
    borde = Border(bottom=Side(style="thin", color=COLOR_BORDE))
    for celda in hoja[1]:
        celda.fill = relleno
        celda.font = Font(color=COLOR_TEXTO, bold=True)
        celda.alignment = Alignment(vertical="center")
        celda.border = borde

    for indice in range(1, hoja.max_column + 1):
        hoja.column_dimensions[get_column_letter(indice)].width = _ancho_columna(
            hoja, indice
        )

    relleno_alterno = PatternFill("solid", fgColor=COLOR_FILA_ALTERNA)
    borde_datos = Border(bottom=Side(style="hair", color=COLOR_BORDE_DATOS))
    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(vertical="top")
            celda.border = borde_datos
            if celda.row % 2 == 0:
                celda.fill = relleno_alterno
            if celda.column_letter and (
                "Importe" in str(hoja.cell(1, celda.column).value)
                or "Diferencia" in str(hoja.cell(1, celda.column).value)
                or "Tolerancia" in str(hoja.cell(1, celda.column).value)
            ):
                if isinstance(celda.value, (int, float)):
                    celda.number_format = '#,##0.00;[Red]-#,##0.00'

    _proteger_texto_formula(hoja)


def _resaltar_diferencias(hoja) -> None:
    relleno_alerta = PatternFill("solid", fgColor=COLOR_ALERTA)
    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.fill = relleno_alerta
            celda.font = Font(color=COLOR_TEXTO)


def exportar_excel(resultado: ResultadoConciliacion) -> bytes:
    """Genera en memoria un XLSX legible y listo para descargar."""

    tablas = {
        "Resumen": resultado.resumen,
        "Coincidencias": resultado.coincidencias,
        "Solo_en_A": resultado.solo_en_a,
        "Solo_en_B": resultado.solo_en_b,
        "Diferencias": resultado.diferencias,
        "Duplicados": resultado.duplicados,
    }
    salida = BytesIO()
    with pd.ExcelWriter(salida, engine="openpyxl") as escritor:
        for nombre, tabla in tablas.items():
            tabla.to_excel(escritor, sheet_name=nombre, index=False)

        libro = escritor.book
        for nombre in NOMBRES_HOJAS:
            hoja = libro[nombre]
            hoja.sheet_properties.tabColor = (
                COLOR_ACENTO if nombre == "Resumen" else COLOR_ENCABEZADO
            )
            _formatear_hoja(hoja)

        _resaltar_diferencias(libro["Diferencias"])
        libro["Resumen"].sheet_properties.pageSetUpPr.fitToPage = True
        libro["Resumen"].sheet_view.zoomScale = 90

    return salida.getvalue()
