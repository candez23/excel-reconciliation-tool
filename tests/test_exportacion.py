from io import BytesIO
from xml.etree import ElementTree
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from conciliador_excel.exportacion import NOMBRES_HOJAS, exportar_excel
from conciliador_excel.reconciliacion import conciliar_archivos


def _resultado_ejemplo():
    datos_a = pd.DataFrame(
        {
            "Clave": ["A-1", "A-2", "A-3", "DUP", "dup"],
            "Importe": [100, 200, 300, 10, 10],
        }
    )
    datos_b = pd.DataFrame(
        {
            "Referencia": ["A-1", "A-2", "A-4", "DUP"],
            "Monto": [100, 205, 400, 10],
        }
    )
    return conciliar_archivos(
        datos_a,
        datos_b,
        "Clave",
        "Referencia",
        "Importe",
        "Monto",
        tolerancia=1,
    )


def test_exporta_las_seis_hojas_con_nombres_exactos_y_formato():
    contenido = exportar_excel(_resultado_ejemplo())
    libro = load_workbook(BytesIO(contenido))

    assert tuple(libro.sheetnames) == NOMBRES_HOJAS
    for nombre in NOMBRES_HOJAS:
        hoja = libro[nombre]
        assert hoja.freeze_panes == "A2"
        assert hoja.sheet_view.showGridLines is False
        assert hoja["A1"].font.bold is True
        assert hoja["A1"].fill.fgColor.rgb.endswith("1D3448")
        color_fila = (
            "6B3940" if nombre == "Diferencias" else "EAF1F5"
        )
        assert hoja["A2"].fill.fgColor.rgb.endswith(color_fila)
        assert len(hoja.tables) == 0
        assert hoja.auto_filter.ref == hoja.dimensions

        columna_inicial, fila_inicial, columna_final, fila_final = range_boundaries(
            hoja.auto_filter.ref
        )
        assert fila_inicial == 1
        assert fila_final >= 2
        assert columna_final >= columna_inicial
    assert libro["Resumen"].max_row > 2
    assert libro["Diferencias"].max_row == 2


def test_ooxml_no_contiene_objetos_ni_relaciones_de_tabla():
    contenido = exportar_excel(_resultado_ejemplo())

    with ZipFile(BytesIO(contenido)) as archivo:
        nombres = archivo.namelist()
        hojas_xml = [
            nombre
            for nombre in nombres
            if nombre.startswith("xl/worksheets/sheet") and nombre.endswith(".xml")
        ]
        relaciones_xml = [
            nombre
            for nombre in nombres
            if nombre.endswith(".rels")
        ]

        assert len(hojas_xml) == len(NOMBRES_HOJAS)
        assert not any(nombre.startswith("xl/tables/") for nombre in nombres)
        assert not any(
            relacion.attrib.get("Type", "").endswith("/table")
            or "/tables/" in relacion.attrib.get("Target", "")
            for nombre in relaciones_xml
            for relacion in ElementTree.fromstring(archivo.read(nombre))
        )


def test_filtros_de_hoja_tienen_el_mismo_rango_valido_que_los_datos():
    contenido = exportar_excel(_resultado_ejemplo())
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    with ZipFile(BytesIO(contenido)) as archivo:
        hojas_xml = [
            nombre
            for nombre in archivo.namelist()
            if nombre.startswith("xl/worksheets/sheet") and nombre.endswith(".xml")
        ]
        for nombre in hojas_xml:
            raiz = ElementTree.fromstring(archivo.read(nombre))
            dimension = raiz.find("x:dimension", namespace)
            filtro = raiz.find("x:autoFilter", namespace)

            assert dimension is not None
            assert filtro is not None
            assert filtro.attrib["ref"] == dimension.attrib["ref"]
            _, fila_inicial, _, fila_final = range_boundaries(filtro.attrib["ref"])
            assert fila_inicial == 1
            assert fila_final >= 2


def test_hojas_sin_registros_no_reciben_un_filtro_vacio():
    datos_a = pd.DataFrame({"Clave": ["A-1"]})
    datos_b = pd.DataFrame({"Referencia": ["A-1"]})
    resultado = conciliar_archivos(datos_a, datos_b, "Clave", "Referencia")

    contenido = exportar_excel(resultado)
    libro = load_workbook(BytesIO(contenido))

    for hoja in libro.worksheets:
        if hoja.max_row == 1:
            assert hoja.auto_filter.ref is None
        else:
            assert hoja.auto_filter.ref == hoja.dimensions
        assert len(hoja.tables) == 0


def test_valores_que_parecen_formula_se_exportan_como_texto():
    datos_a = pd.DataFrame({"Clave": ["=2+2"], "Valor": ["=CMD()"]})
    datos_b = pd.DataFrame({"Referencia": ["=2+2"], "Dato": ["seguro"]})
    resultado = conciliar_archivos(datos_a, datos_b, "Clave", "Referencia")

    contenido = exportar_excel(resultado)
    libro = load_workbook(BytesIO(contenido), data_only=False)
    hoja = libro["Coincidencias"]
    encabezados = {
        celda.value: celda.column for celda in hoja[1] if celda.value is not None
    }
    celda_clave = hoja.cell(2, encabezados["A | Clave"])
    celda_valor = hoja.cell(2, encabezados["A | Valor"])

    assert celda_clave.value == "=2+2"
    assert celda_clave.data_type == "s"
    assert celda_valor.value == "=CMD()"
    assert celda_valor.data_type == "s"
